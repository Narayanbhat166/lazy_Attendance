from PIL import Image
from openpyxl import Workbook,load_workbook
import pytesseract
import glob
import cv2
import os
import time
import json
import datetime
import shutil
import requests

bot_url = 'https://api.telegram.org/bot1291011387:AAEDG2wqE0t4XHbe_9RurkcJHJ_Fdw99rf8/'


def Send_Document(chat_id):
    x = datetime.datetime.now()
    FILE = '/home/chaser/telegram/python-bot/Database/Saved_Files/'
    FILE += x.strftime("%B")+x.strftime("%d")+".xlsx"

    method = 'sendDocument'

    os.chdir('/home/chaser/telegram/python-bot/Database/Saved_Files')

    document = open(FILE,'rb')
    response = requests.post(bot_url + method, data={'chat_id': chat_id}, files={'document': document})
    document.close()

    response = response.text

    with open('FileHistory.json','a') as f:
        f.write(response)

def reply(message,chat_id):

    json_data = {
        "chat_id": chat_id,
        "text":message
    }

    message_url = bot_url+'sendMessage'
    requests.post(message_url, json=json_data)
# To keep track of the cell numbers in excel(You dont need to worry about that)

class excel():
    def __init__(self):
        x = datetime.datetime.now()
        self.FILE_NAME = '/home/chaser/telegram/python-bot/Database/Saved_Files/'
        self.FILE_NAME += x.strftime("%B")+x.strftime("%d")+".xlsx"
        self.cell_no = 1
        try:
            self.wb = load_workbook(filename=self.FILE_NAME)
            print("Opened exixsing workbook")
        except:
            self.wb = Workbook()
            print("opened new workbook")
        self.ws = self.wb.active
    

    def Insert(self,Name):
        if not Name.__eq__(' '):
            content = self.ws['A'+str(self.cell_no)].value
            if content:
                points =  self.ws['B'+str(self.cell_no)].value
                points = points + 1
                self.ws['B'+str(self.cell_no)] = points
            else:
                self.ws['A'+str(self.cell_no)] = Name
                self.ws['B'+str(self.cell_no)] = 1
        self.cell_no = self.cell_no+1
    
    def Save(self):
        self.wb.save(self.FILE_NAME)
        try:
            self.wb.close()
        except:
            print("Error occured when closing the excel file")

def Extract_Text(sender):
    xl = excel();
    # All the students in the class for comparing
    Students = {}

    FILE_NAME = "Students.json"
    os.chdir("/home/chaser/telegram/python-bot/Database")
    with open(FILE_NAME, 'r') as f:
        Students = json.load(f)

    # if error "tesseract not found", then add this line including the path of the installed module
    pytesseract.pytesseract.tesseract_cmd = "/usr/bin/tesseract"

    # List of students present
    present = []
    # To scan the required frames
    frame_no = 0
    increment_value = 5

    #Extract the images one by one, from the path specified
    while True:
        try:
        # Location of the image(with name)
            image = f"/home/chaser/telegram/python-bot/Database/Extracted_frames/frame{frame_no}.jpg"
            print(sender+":Scanning "+image)
            img = Image.open(image)
            width, height = img.size

            croppedimage = img.crop((0, 300, width, height))
            text = pytesseract.image_to_string(croppedimage)
            text = text.lower()
            for student in Students:
                if student.lower() in text:
                    if student not in present:
                        present.append(student)

            frame_no += increment_value
        except:
            print(sender+"Completed Scan")
            # if no more frames available to read, break
            break

    #Insert the student details into the excel sheet(As per the roll number), if absent then insert " "(space)
    for student in Students:
        if student in present:
            xl.Insert(student)
        else:
            xl.Insert(" ")

   
    xl.Save()
    print(present)
    print("Students present Today = "+str(len(present)))
    global cell_no
    cell_no = 1

def Create_Frames(file,sender,chat_id,lock):
    lock.acquire()
    reply("Processing Your Video, Relax!",chat_id)

    # Video must be names as Attendance.mp4
    video = cv2.VideoCapture(f"/home/chaser/telegram/python-bot/Downloaded_Files/{file}")

    # Directory name to save the extracted images
    dirs = "/home/chaser/telegram/python-bot/Database/Extracted_frames"

    # Create an empty directory even if it exists(Overwrite)
    try:
        if os.path.exists(dirs):
            shutil.rmtree(dirs)
        os.makedirs(dirs)
    except OSError:
        print("Error creating directory")

    # Have a count of frame number to save the frames
    currentframe = 0
    os.chdir(f"/home/chaser/telegram/python-bot/Database/Extracted_frames/")

    # Extract all the frames from the vide0
    while True:
        ret, frame = video.read()

        if ret:
            name = 'frame'+str(currentframe)+'.jpg'
            print("Creating...."+name)
            cv2.imwrite(name, frame)
            currentframe += 1
        else:
            break

    # Release the resources
    video.release()
    print("Created "+str(currentframe)+" Frames in directory"+dirs)

    Extract_Text(sender)


    lock.release()
    reply("Document Ready!",chat_id)
    Send_Document(chat_id)






